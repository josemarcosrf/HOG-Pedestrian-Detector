# -*- coding: utf8 -*-

import os
import copy
import string

# openpyxl imports
from openpyxl.style import Color, Fill, Alignment, Border
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import get_column_letter




class workbook(object):
    
    Color.myGreen = '7EBA4C'
    Color.myRed = 'C7141D'
    Color.myGray = 'C4C4C4'
    rows2add = []
    cols2add = []


    """ BOOK STAFF """
    def init_book(self):
        # creamos el libro de excel
        self._wb = Workbook()
        return self._wb.worksheets[0]

    def get_book(self):
        return self._wb
        
    def set_book(self, wb):
        self._wb = wb
        
    def save_book(self,path,name):
        final_path = os.path.join(path, name+'.xlsx')
        if not os.path.exists(path):
            print "creating path", path
            os.makedirs(path)
        
        print "saving workbook as:", final_path
        self._wb.save(final_path)
        
    
    """ COLUMNS AND ROWS ADDING STAFF """
    def set_rows2add(self,list):
        self.rows2add = copy.deepcopy(list)
        
    def set_cols2add(self,list):
        self.cols2add = copy.deepcopy(list)

    def get_rows2add(self):
        return self.rows2add
        
    def get_cols2add(self):
        return self.cols2add
        
      
    """  SHEET STAFF     """
    def get_sheet(self, i):
        try:
            return self._wb.worksheets[i]
        except:
            print "ERROR: no sheet with that number..."
            
    def set_sheet(self, ws, i):
        try:
            self._wb.worksheets[i] = ws
        except:
            print "Not possible to set sheet"+str(i)+"..."
            
    def get_sheet_title(self, i):
        try:
            return self._wb[i].title
        except:
            print "Not possible to retrieve sheet"+str(i)+"title..."
    
    def get_sheet_count(self):
        return len(self._wb.worksheets)
        
    def get_sheet_titles(self):
        return [s.title for s in self._wb.worksheets]

    def create_new_sheet(self, title):
        ws = self._wb.create_sheet()
        ws.title = title
        return self._wb.worksheets[-1]
        
    def set_sheet_title(self,sheet_num, s_title):
        try:
            self._wb.worksheets[sheet_num].title = s_title
        except:
            print "ERROR: Cannot access that sheet number..."
        
    def set_sheet_title(self,ws, s_title):
        ws.title = s_title
        
    


    """ TABLE STAFF """
    def write_table(self,ws,table, v_titles=None, h_titles=None, v_sum=False, h_sum=False, d_sum=False,highlight_diagonal=False, highlight_max=False):

        filas = len(table)
        columnas = len(table[0])
        
        # creamos copias de todo porque Python pasa todo por referencia
        _table = copy.deepcopy(table)
        _h_titles = copy.deepcopy(h_titles)
        _v_titles = copy.deepcopy(v_titles)
        
        # calculo del valor máximo de la tabla
        if highlight_max:
            maximum = max(max(l) for l in _table)
            
        # calculo de la suma de la diagonal
        if d_sum:
            table_short_side = min(len(_table), len(_table[0]))
            d_suma = sum([_table[i][i] for i in range( table_short_side )])
          
        # informacion adicional (sumas filas, columnas, diagonales...)
        filas, columnas = self.__aditional_inf(_table, _v_titles, _h_titles, filas, columnas,v_sum, h_sum)
        
        # escritura de titulos de tabla (si existen)
        h_desp, v_desp = self.__write_titles(ws, _v_titles, _h_titles)
        
        # indices verticales de tabla
        ind_verticales = [str(i+v_desp) for i in range(filas)]
        
        #escritura de tabla
        for f in range(filas):
            for c in range(columnas):
                
                # if c == columnas-1 and f == filas-1:
                #     continue
                
                col = get_column_letter(c+h_desp)
                cell = ws.cell(col+ind_verticales[f])
                cell.value = _table[f][c]
                cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
                
                # maximo de la tabla
                if highlight_max and _table[f][c] == maximum:
                    cell.style.fill.fill_type = Fill.FILL_SOLID
                    cell.style.fill.start_color.index = Color.myRed
                
                # diagonal de la tabla
                if highlight_diagonal and c == f: 
                    if not (c == columnas-1 and h_sum) and \
                       not (f == filas-1 and v_sum):
                        cell.style.fill.fill_type = Fill.FILL_SOLID
                        cell.style.fill.start_color.index = Color.myGreen
                
                # suma de columnas
                if f == filas-1 and h_sum and len(self.cols2add) > 0:
                    cell.style.font.bold = True
                    if c not in self.cols2add:
                        cell.value = ""
                    
                # suma de filas 
                if c == columnas-1 and v_sum and len(self.rows2add) > 0:
                    cell.style.font.bold = True
                    if f not in self.rows2add:
                        cell.value = ""
            
        # escritura de la suma de la diagonal
        if d_sum:
            col = get_column_letter(columnas+h_desp-1)
            row = str(filas+v_desp-1)
            if not (h_sum and len(self.rows2add) > 0): col = get_column_letter(columnas+h_desp) 
            if not (v_sum and len(self.cols2add) > 0): row = str(int(row)+1)
            
            cell = ws.cell(col+row)
            cell.value = d_suma
            cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
            cell.style.fill.fill_type = Fill.FILL_SOLID
            cell.style.fill.start_color.index = Color.myGreen
            cell.style.font.bold = True
                    
        
                
        
    """
    ////////////////////////////////////////////////////////////////////////////////////
                                FUNCIONES AUXILIARES Y PRIVADAS
    ////////////////////////////////////////////////////////////////////////////////////
    """
    
    def __write_titles(self,ws, v_titles, h_titles):
        h_desp = v_desp = 1
        # --titulos verticales--
        if v_titles != None: 
            v_desp = h_desp = 2
            for t in range(len(v_titles)):
                ind = str(t+2)
                cell = ws.cell('A'+str(ind))
                cell.value = v_titles[t]
                cell.style.font.bold = True
                cell.style.alignment.wrap_text = True
                # cell.style.fill.fill_type = Fill.FILL_PATTERN_LIGHTGRAY
                cell.style.fill.fill_type = Fill.FILL_SOLID
                cell.style.fill.start_color.index = Color.myGray
                cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
                cell.style.border = Border.BORDER_DOUBLE
        
        # --titulos horizontales--
        if h_titles != None: 
            h_desp = v_dsp = 2
            for h in range(len(h_titles)):
                ind = h+2
                col = get_column_letter(ind)
                cell = ws.cell(col+'1')
                cell.value = h_titles[h]
                cell.style.font.bold = True
                cell.style.alignment.wrap_text = True
                # cell.style.fill.fill_type = Fill.FILL_PATTERN_LIGHTGRAY
                cell.style.fill.fill_type = Fill.FILL_SOLID
                cell.style.fill.start_color.index = Color.myGray
                cell.style.alignment.wrap_text = True
                cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
                cell.style.border = Border.BORDER_DOUBLE
                
        return h_desp, v_desp
        
        
    def __aditional_inf(self,_table, v_titles, h_titles, filas, columnas,v_sum, h_sum):

        # corrección de títulos si hiciese falta
        while len(v_titles) < filas:
            v_titles.append("N/D")
            
        while len(h_titles) < columnas:
            h_titles.append("N/D")
            
        # suma de los valores de las filas
        if h_sum and len(self.rows2add) > 0:
            for fila in range(filas):
                suma = 0
                if fila in self.rows2add:
                    suma = sum(_table[fila][:])
                
                _table[fila].append(suma)
                
        
        # suma de los valores de las columnas 
        if v_sum and len(self.cols2add) > 0:
            totales = []
            
            for c in range(columnas):
                if c in self.cols2add:
                    total = 0
                    for f in range(filas):
                        total += _table[f][c]
                        
                    totales.append(total)
            
            # nueva fila con los totales por columna
            # (haciendo la ultima fila de la misma longitud que el resto)
            while(len(totales) < len(_table[-1])):
                totales.append("")

            _table.append(totales)
        

        # actualizando numero de filas y columnas
        if v_sum and len(self.cols2add) > 0:
            filas += 1
            v_titles.append("total")
            
        if h_sum and len(self.rows2add) > 0:
            columnas += 1
            h_titles.append("total")

        return filas, columnas
        
    
    
    
def main():
    # creamos una table cualquiera
    table = [[i*(j+1) for i in range(10)] for j in range(10)]

    wb = workbook()
    ws1 = wb.init_book()
   
    # primer hoja del libro
    wb.set_sheet_title(ws1,"primera pagina")
    
    # creamos una nueva pagina
    print "creando pagina 2"
    ws2 = wb.create_new_sheet("pagina 2")
    
    # contamos cuantas paginas
    print "\ncantidad de paginas:", wb.get_sheet_count()
    print "titutlos:"
    for p in wb.get_sheet_titles():
        print "* ",p
        

    # wb.set_cols2add(range(len(table)))
    wb.set_rows2add([0,1,2])
    wb.write_table(ws1,table,h_titles=[str(i)+'*' for i in range(10)],v_titles=['a','b','c','d'], v_sum=True,d_sum=True,h_sum=True,highlight_diagonal=True)
    wb.write_table(ws2,table,h_titles=[str(i)+'*' for i in range(10)],v_titles=['a','b','c','d'], v_sum=False,highlight_diagonal=True)

    wb.save_book('.','prueba')
    # raw_input("...")

    
if __name__=="__main__":
    main()



